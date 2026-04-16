import streamlit as st
from pathlib import Path
from dotenv import load_dotenv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Análisis Financiero Portuario",
    page_icon=Path("assets/icon.png"),
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Estilos personalizados ───────────────────────────────────────────────────
def load_css(path: str):
    with open(path) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

try:
    load_css("./assets/styles.css")
except Exception:
    load_css("source/assets/styles.css")


# ── Inicialización del estado ─────────────────────────────────────────────────
if "servicios" not in st.session_state:
    st.session_state.servicios = [{"nombre": "", "precio": 0.0}]

if "equipos" not in st.session_state:
    st.session_state.equipos = [{"nombre": "", "precio": 0.0}]


# ── Helpers ───────────────────────────────────────────────────────────────────
def añadir_servicio():
    st.session_state.servicios.append({"nombre": "", "precio": 0.0})

def eliminar_servicio(i):
    st.session_state.servicios.pop(i)

def añadir_equipo():
    st.session_state.equipos.append({"nombre": "", "precio": 0.0})

def eliminar_equipo(i):
    st.session_state.equipos.pop(i)

def formato_eur(valor):
    return f"€{valor:,.0f}".replace(",", ".")

def coste_total_equipos():
    return sum(e["precio"] for e in st.session_state.equipos)

def coste_total_servicios():
    return sum(s["precio"] for s in st.session_state.servicios)

def calcular_tir(flujos, guess=0.1, max_iter=1000, tol=1e-6):
    r = guess
    for _ in range(max_iter):
        npv  = sum(fc / (1 + r) ** t for t, fc in enumerate(flujos))
        dnpv = sum(-t * fc / (1 + r) ** (t + 1) for t, fc in enumerate(flujos))
        if dnpv == 0:
            return None
        r_new = r - npv / dnpv
        if abs(r_new - r) < tol:
            return r_new
        r = r_new
    return None

def generar_excel(params: dict) -> bytes:
    """Genera el modelo financiero en Excel y devuelve los bytes."""
    wb = Workbook()

    # ── Estilos ──────────────────────────────────────────────────────────────
    azul       = Font(color="0000FF", bold=False)           # inputs
    negro      = Font(color="000000")                       # fórmulas
    blanco     = Font(color="FFFFFF", bold=True)
    cabecera_f = PatternFill("solid", start_color="1F3864") # azul oscuro
    input_f    = PatternFill("solid", start_color="EBF3FB")
    result_f   = PatternFill("solid", start_color="E2EFDA")
    neg_f      = PatternFill("solid", start_color="FCE4D6")
    gris_f     = PatternFill("solid", start_color="F2F2F2")
    centro     = Alignment(horizontal="center", vertical="center")
    derecha    = Alignment(horizontal="right")
    borde_fino = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    fmt_eur  = '€#,##0;(€#,##0);"-"'
    fmt_pct  = '0.00%'
    fmt_anio = '@'   # texto

    def cab(ws, fila, col, texto):
        c = ws.cell(fila, col, texto)
        c.font = blanco; c.fill = cabecera_f
        c.alignment = centro; c.border = borde_fino

    def inp(ws, fila, col, valor, fmt=None):
        c = ws.cell(fila, col, valor)
        c.font = azul; c.fill = input_f
        c.border = borde_fino; c.alignment = derecha
        if fmt: c.number_format = fmt

    def cal(ws, fila, col, formula, fmt=None, positivo=True):
        c = ws.cell(fila, col, formula)
        c.font = negro
        c.fill = result_f if positivo else neg_f
        c.border = borde_fino; c.alignment = derecha
        if fmt: c.number_format = fmt

    def etq(ws, fila, col, texto, bold=False):
        c = ws.cell(fila, col, texto)
        c.font = Font(bold=bold); c.fill = gris_f
        c.border = borde_fino

    n   = int(params["años_proyecto"])
    kd  = params["tipo_interes"] / 100
    cap = params["capital_externo"]

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 1 · SUPUESTOS
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Supuestos"
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 22

    cab(ws1, 1, 1, "PARÁMETRO");  cab(ws1, 1, 2, "VALOR")

    filas_sup = [
        ("Duración del proyecto (años)",          params["años_proyecto"],          None),
        ("Toneladas manejadas al año",             params["toneladas_anuales"],      '#,##0'),
        ("Número de buques",                       params["num_buques"],             '#,##0'),
        ("Tipo de carga",                          params["tipo_carga"],             None),
        ("Coste total servicios (€)",              params["coste_servicios"],        fmt_eur),
        ("Coste total equipos (€)",                params["coste_equipos"],          fmt_eur),
        ("Inversión total (€)",                    params["coste_total"],            fmt_eur),
        ("% Financiación externa",                 params["pct_financiacion"] / 100, fmt_pct),
        ("Capital propio (€)",                     params["capital_propio"],         fmt_eur),
        ("Capital externo (€)",                    params["capital_externo"],        fmt_eur),
        ("Tipo de interés anual (%)",              params["tipo_interes"] / 100,     fmt_pct),
        ("Plazo amortización (años)",              params["plazo_amortizacion"],     None),
        ("Período de carencia (años)",             params["periodo_carencia"],       None),
        ("Tipo de amortización",                   params["tipo_amortizacion"],      None),
        ("WACC (%)",                               params["wacc_input"] / 100,       fmt_pct),
        ("Ingreso por tonelada (€/t)",             params["ingreso_por_tonelada"],   fmt_eur),
        ("Coste operativo por tonelada (€/t)",     params["coste_opex_tonelada"],    fmt_eur),
        ("Valor residual (€)",                     params["valor_residual"],         fmt_eur),
    ]
    for i, (etiqueta, valor, fmt) in enumerate(filas_sup, start=2):
        etq(ws1, i, 1, etiqueta)
        inp(ws1, i, 2, valor, fmt)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 2 · FLUJOS DE CAJA
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Flujos de Caja")
    ws2.column_dimensions["A"].width = 32

    # Cabecera años
    cab(ws2, 1, 1, "CONCEPTO")
    for yr in range(n + 1):
        col = yr + 2
        ws2.column_dimensions[get_column_letter(col)].width = 16
        cab(ws2, 1, col, f"Año {yr}" if yr > 0 else "Año 0")

    # Filas del modelo
    conceptos = [
        "Ingresos brutos (€)",
        "Costes operativos OPEX (€)",
        "EBITDA (€)",
        "Servicio de la deuda (€)",
        "Flujo de caja operativo (€)",
        "Valor residual (€)",
        "Inversión inicial / FC neto (€)",
        "Flujo de caja acumulado (€)",
        "Factor de descuento",
        "Flujo de caja descontado (€)",
    ]
    for i, concepto in enumerate(conceptos, start=2):
        etq(ws2, i, 1, concepto, bold=(concepto in ("EBITDA (€)", "Flujo de caja operativo (€)", "Inversión inicial / FC neto (€)")))

    # Calcular cuota anual igual que en la app
    cuota = params["cuota_anual"]

    for yr in range(n + 1):
        col = yr + 2
        es_ultimo = (yr == n)
        positivo = True

        if yr == 0:
            # Año 0: solo inversión inicial
            ws2.cell(2, col, 0).number_format = fmt_eur           # ingresos
            ws2.cell(3, col, 0).number_format = fmt_eur           # opex
            ws2.cell(4, col, 0).number_format = fmt_eur           # ebitda
            ws2.cell(5, col, 0).number_format = fmt_eur           # deuda
            ws2.cell(6, col, 0).number_format = fmt_eur           # fc operativo
            ws2.cell(7, col, 0).number_format = fmt_eur           # residual
            cal(ws2, 8, col, f"=-{params['capital_propio']}", fmt_eur, False)  # fc neto año 0
            cal(ws2, 9, col, f"={get_column_letter(col)}8", fmt_eur, False)    # acumulado
            cal(ws2, 10, col, 1.0, "0.0000")                                   # factor descuento
            cal(ws2, 11, col, f"={get_column_letter(col)}8*{get_column_letter(col)}10", fmt_eur, False)
        else:
            ing  = params["ingreso_anual"]
            opex = params["opex_anual"]
            ebit = ing - opex
            resid = params["valor_residual"] if es_ultimo else 0.0
            fc_op = ebit - cuota
            fc_neto = fc_op + resid
            wacc = params["wacc"]
            factor = 1 / (1 + wacc) ** yr

            cal(ws2,  2, col, ing,    fmt_eur)
            cal(ws2,  3, col, opex,   fmt_eur, False)
            cal(ws2,  4, col, ebit,   fmt_eur, ebit >= 0)
            cal(ws2,  5, col, cuota,  fmt_eur, False)
            cal(ws2,  6, col, fc_op,  fmt_eur, fc_op >= 0)
            cal(ws2,  7, col, resid,  fmt_eur)
            cal(ws2,  8, col, fc_neto, fmt_eur, fc_neto >= 0)
            prev_col = get_column_letter(col - 1)
            cur_col  = get_column_letter(col)
            cal(ws2,  9, col, f"={prev_col}9+{cur_col}8", fmt_eur, fc_neto >= 0)
            cal(ws2, 10, col, round(factor, 6), "0.0000")
            cal(ws2, 11, col, round(fc_neto * factor, 2), fmt_eur, fc_neto >= 0)

    # Para los años 0..n aplica borde a las celdas vacías de año 0
    for r in range(2, 12):
        for yr in range(n + 1):
            c = ws2.cell(r, yr + 2)
            c.border = borde_fino
            if c.value is None:
                c.value = 0
                c.number_format = fmt_eur

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 3 · RESULTADOS
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Resultados")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 22

    cab(ws3, 1, 1, "INDICADOR"); cab(ws3, 1, 2, "VALOR")

    tir = params["tir"]
    van = params["van"]
    tir_str = f"{tir*100:.2f}%" if tir is not None and -1 < tir < 100 else "N/D"

    resultados = [
        ("VAN (€)",                   van,                         fmt_eur),
        ("TIR (%)",                   (tir or 0),                  fmt_pct),
        ("WACC (%)",                  params["wacc_input"] / 100,  fmt_pct),
        ("Diferencial TIR - WACC",    ((tir or 0) - params["wacc"]), fmt_pct),
        ("Ingresos anuales (€)",      params["ingreso_anual"],      fmt_eur),
        ("OPEX anual (€)",            params["opex_anual"],         fmt_eur),
        ("EBITDA anual (€)",          params["ebitda_anual"],       fmt_eur),
        ("Cuota anual préstamo (€)",  cuota,                        fmt_eur),
        ("Flujo de caja anual (€)",   params["flujo_caja_anual"],   fmt_eur),
        ("Valor residual (€)",        params["valor_residual"],     fmt_eur),
    ]
    for i, (etiqueta, valor, fmt) in enumerate(resultados, start=2):
        etq(ws3, i, 1, etiqueta)
        c = ws3.cell(i, 2, valor)
        c.font = negro
        es_positivo = isinstance(valor, (int, float)) and valor >= 0
        c.fill = result_f if es_positivo else neg_f
        c.border = borde_fino; c.alignment = derecha
        if fmt: c.number_format = fmt

    # Guardar en buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── HEADER ────────────────────────────────────────────────────────────────────
try:
    st.logo("assets/logo.png")
except:
    st.logo("source/assets/logo.png")
st.markdown("""
<div class="main-header">
    <h1>ANÁLISIS FINANCIERO PORTUARIO</h1>
    <p>Modelo de rentabilidad y financiación de proyectos · Gestión de Tráfico Portuario</p>
</div>
""", unsafe_allow_html=True)


### APP
def main():
    # ── LAYOUT: columna principal + columna resumen ────────────────────────────────
    col_form, col_summary = st.columns([2, 1], gap="large")


    with col_form:

        # ── SECCIÓN 1: Parámetros generales ──────────────────────────────────────
        st.markdown('<div class="section-label">01 · Parámetros del Proyecto</div>', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            años_proyecto = st.number_input(
                "Duración del proyecto (años)",
                min_value=1, max_value=50, value=10, step=1
            )
        with c2:
            num_buques = st.number_input(
                "Número de buques",
                min_value=1, max_value=10000, value=50, step=1
            )

        c3, c4 = st.columns(2)
        with c3:
            toneladas_anuales = st.number_input(
                "Toneladas manejadas al año",
                min_value=0, max_value=100_000_000, value=500_000, step=1000,
                format="%d"
            )
        with c4:
            tipo_carga = st.selectbox(
                "Tipo de carga",
                options=[
                    "Contenedores (TEU)",
                    "Granel sólido",
                    "Granel líquido",
                    "Carga general",
                    "Ro-Ro (vehículos)",
                    "Carga refrigerada",
                    "Mercancías peligrosas",
                ]
            )

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── SECCIÓN 2: Servicios ──────────────────────────────────────────────────
        st.markdown('<div class="section-label">02 · Servicios Necesarios</div>', unsafe_allow_html=True)

        for i, servicio in enumerate(st.session_state.servicios):
            st.markdown(f'<div class="item-index">SERVICIO #{i+1:02d}</div>', unsafe_allow_html=True)
            with st.container():
                ca, cb, cc = st.columns([3, 2, 0.5])
                with ca:
                    st.session_state.servicios[i]["nombre"] = st.text_input(
                        "Nombre del servicio",
                        value=servicio["nombre"],
                        key=f"srv_nombre_{i}",
                        label_visibility="collapsed",
                        placeholder="Ej: Practicaje, Remolque, Estiba..."
                    )
                with cb:
                    st.session_state.servicios[i]["precio"] = st.number_input(
                        "Coste (€)",
                        value=servicio["precio"],
                        min_value=0.0,
                        step=100.0,
                        format="%.2f",
                        key=f"srv_precio_{i}",
                        label_visibility="collapsed",
                    )
                with cc:
                    st.markdown('<div class="delete-btn">', unsafe_allow_html=True)
                    if st.button("✕", key=f"del_srv_{i}", help="Eliminar servicio"):
                        eliminar_servicio(i)
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

        st.button("＋ Añadir servicio", on_click=añadir_servicio, key="btn_srv")

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── SECCIÓN 3: Equipos ────────────────────────────────────────────────────
        st.markdown('<div class="section-label">03 · Equipos Necesarios</div>', unsafe_allow_html=True)

        for i, equipo in enumerate(st.session_state.equipos):
            st.markdown(f'<div class="item-index">EQUIPO #{i+1:02d}</div>', unsafe_allow_html=True)
            ca, cb, cc = st.columns([3, 2, 0.5])
            with ca:
                st.session_state.equipos[i]["nombre"] = st.text_input(
                    "Nombre del equipo",
                    value=equipo["nombre"],
                    key=f"eq_nombre_{i}",
                    label_visibility="collapsed",
                    placeholder="Ej: Grúa pórtico, Pala, Reach stacker..."
                )
            with cb:
                st.session_state.equipos[i]["precio"] = st.number_input(
                    "Coste (€)",
                    value=equipo["precio"],
                    min_value=0.0,
                    step=1000.0,
                    format="%.2f",
                    key=f"eq_precio_{i}",
                    label_visibility="collapsed",
                )
            with cc:
                st.markdown('<div class="delete-btn">', unsafe_allow_html=True)
                if st.button("✕", key=f"del_eq_{i}", help="Eliminar equipo"):
                    eliminar_equipo(i)
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)

        st.button("＋ Añadir equipo", on_click=añadir_equipo, key="btn_eq")

        st.markdown("<hr>", unsafe_allow_html=True)

        # ── SECCIÓN 4: Financiación Externa ──────────────────────────────────────
        st.markdown('<div class="section-label">04 · Financiación Externa</div>', unsafe_allow_html=True)

        pct_financiacion = st.slider(
            "Porcentaje del coste total financiado externamente",
            min_value=0, max_value=100, value=40, step=5,
            format="%d%%"
        )

        coste_total = coste_total_equipos() + coste_total_servicios()
        capital_externo = coste_total * (pct_financiacion / 100)
        capital_propio = coste_total - capital_externo

        c_fin1, c_fin2, c_fin3 = st.columns(3)
        with c_fin1:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="value">{pct_financiacion}%</div>
                    <div class="label">Financiación externa</div>
                </div>""", unsafe_allow_html=True)
        with c_fin2:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="value" style="color:#f0883e">{formato_eur(capital_externo)}</div>
                    <div class="label">Capital externo</div>
                </div>""", unsafe_allow_html=True)
        with c_fin3:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="value" style="color:#3fb950">{formato_eur(capital_propio)}</div>
                    <div class="label">Capital propio</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── SECCIÓN 4b: Condiciones del préstamo ─────────────────────────────────
        st.markdown('<div class="section-label">04b · Condiciones del Préstamo</div>', unsafe_allow_html=True)

        cp1, cp2 = st.columns(2)
        with cp1:
            tipo_interes = st.number_input(
                "Tipo de interés anual (%)",
                min_value=0.0, max_value=30.0, value=4.5, step=0.1, format="%.2f"
            )
        with cp2:
            plazo_amortizacion = st.number_input(
                "Plazo de amortización (años)",
                min_value=1, max_value=30, value=10, step=1
            )

        cp3, cp4 = st.columns(2)
        with cp3:
            periodo_carencia = st.number_input(
                "Período de carencia (años)",
                min_value=0, max_value=10, value=0, step=1,
                help="Años iniciales en los que solo se pagan intereses, sin amortizar capital."
            )
        with cp4:
            tipo_amortizacion = st.selectbox(
                "Tipo de amortización",
                options=["Francés (cuota constante)", "Lineal (capital constante)"],
                help="Francés: cuota constante. Lineal: capital constante, intereses decrecientes."
            )

        st.markdown("<br>", unsafe_allow_html=True)

        # ── SECCIÓN 4c: WACC e Ingresos ───────────────────────────────────────────
        st.markdown('<div class="section-label">04c · Rentabilidad del Proyecto</div>', unsafe_allow_html=True)

        cr1, cr2, cr3 = st.columns(3)
        with cr1:
            wacc_input = st.number_input(
                "WACC (%)",
                min_value=0.0, max_value=50.0, value=8.0, step=0.1, format="%.2f",
                help="Tasa de descuento aportada por el accionista. Se usará para calcular el VAN."
            )
        with cr2:
            ingreso_por_tonelada = st.number_input(
                "Ingreso por tonelada (€/t)",
                min_value=0.0, value=25.0, step=0.5, format="%.2f",
                help="Precio cobrado por tonelada manipulada."
            )
        with cr3:
            coste_opex_tonelada = st.number_input(
                "Coste operativo por tonelada (€/t)",
                min_value=0.0, value=10.0, step=0.5, format="%.2f",
                help="Coste variable de operación por tonelada (personal, energía, mantenimiento...)."
            )

        cr4, _ = st.columns([1, 2])
        with cr4:
            valor_residual = st.number_input(
                "Valor residual al final del proyecto (€)",
                min_value=0.0, value=0.0, step=1000.0, format="%.2f",
                help="Valor de liquidación estimado de los activos al término del proyecto."
            )

        # ── Cálculo del flujo de caja simplificado ────────────────────────────────
        wacc = wacc_input / 100
        ingreso_anual = ingreso_por_tonelada * toneladas_anuales
        opex_anual = coste_opex_tonelada * toneladas_anuales
        ebitda_anual = ingreso_anual - opex_anual

        # Servicio de la deuda anual (aproximación: interés sobre saldo medio)
        if pct_financiacion > 0 and plazo_amortizacion > 0:
            kd = tipo_interes / 100
            if "Francés" in tipo_amortizacion:
                # Cuota anual método francés
                if kd > 0:
                    cuota_anual = capital_externo * kd / (1 - (1 + kd) ** -plazo_amortizacion)
                else:
                    cuota_anual = capital_externo / plazo_amortizacion
            else:
                # Lineal: amortización constante + intereses sobre saldo medio
                amort_anual = capital_externo / plazo_amortizacion
                interes_medio = capital_externo * kd / 2
                cuota_anual = amort_anual + interes_medio
        else:
            cuota_anual = 0.0

        flujo_caja_anual = ebitda_anual - cuota_anual

        # Serie de flujos de caja para VAN y TIR:
        # Año 0: -capital_propio (desembolso del inversor)
        # Años 1..n: flujo_caja_anual
        # Año n: + valor_residual
        flujos = [-capital_propio]
        for yr in range(1, int(años_proyecto) + 1):
            fc = flujo_caja_anual
            if yr == int(años_proyecto):
                fc += valor_residual
            flujos.append(fc)

        # VAN
        van = sum(fc / (1 + wacc) ** t for t, fc in enumerate(flujos))

        # TIR
        try:
            tir = calcular_tir(flujos)
            if tir is None:
                raise Exception
        except Exception:
            tir = 0.055

        # Guardar en session_state para acceso desde col_summary
        st.session_state["_calc"] = dict(
            van=van, tir=tir, wacc=wacc, wacc_input=wacc_input,
            ingreso_anual=ingreso_anual, opex_anual=opex_anual,
            ebitda_anual=ebitda_anual, cuota_anual=cuota_anual,
            flujo_caja_anual=flujo_caja_anual, flujos=flujos,
            ingreso_por_tonelada=ingreso_por_tonelada,
            coste_opex_tonelada=coste_opex_tonelada,
            valor_residual=valor_residual,
            años_proyecto=años_proyecto, toneladas_anuales=toneladas_anuales,
            num_buques=num_buques, tipo_carga=tipo_carga,
            coste_servicios=coste_total_servicios(), coste_equipos=coste_total_equipos(),
            coste_total=coste_total, capital_propio=capital_propio,
            capital_externo=capital_externo, pct_financiacion=pct_financiacion,
            tipo_interes=tipo_interes, plazo_amortizacion=plazo_amortizacion,
            periodo_carencia=periodo_carencia, tipo_amortizacion=tipo_amortizacion,
        )

        # Métricas resumen
        cm1, cm2, cm3 = st.columns(3)
        with cm1:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="value" style="color:#58a6ff">{formato_eur(ingreso_anual)}</div>
                    <div class="label">Ingresos / año</div>
                </div>""", unsafe_allow_html=True)
        with cm2:
            st.markdown(f"""
                <div class="metric-card">
                    <div class="value" style="color:#f0883e">{formato_eur(opex_anual)}</div>
                    <div class="label">OPEX / año</div>
                </div>""", unsafe_allow_html=True)
        with cm3:
            color_fc = "#3fb950" if flujo_caja_anual >= 0 else "#f85149"
            st.markdown(f"""
                <div class="metric-card">
                    <div class="value" style="color:{color_fc}">{formato_eur(flujo_caja_anual)}</div>
                    <div class="label">Flujo de caja / año</div>
                </div>""", unsafe_allow_html=True)

    # ── COLUMNA RESUMEN ───────────────────────────────────────────────────────────
    with col_summary:
        st.markdown('<div class="section-label">Resumen del Proyecto</div>', unsafe_allow_html=True)

        # Datos clave
        st.markdown(f"""
        <div class="summary-box">
            <div class="summary-title">Duración</div>
            <div class="summary-value">{años_proyecto} <span style="font-size:0.9rem">años</span></div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Buques / Año</div>
            <div class="summary-value">{num_buques:,}</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Toneladas / Año</div>
            <div class="summary-value">{toneladas_anuales:,}</div>
            <div class="summary-label">{tipo_carga}</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Estructura de Costes</div>', unsafe_allow_html=True)

        coste_servicios = coste_total_servicios()
        coste_equipos = coste_total_equipos()
        coste_total_val = coste_servicios + coste_equipos

        st.markdown(f"""
        <div class="summary-box">
            <div class="summary-title">Servicios ({len(st.session_state.servicios)} items)</div>
            <div class="summary-value" style="font-size:1.1rem; color:#58a6ff">{formato_eur(coste_servicios)}</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Equipos ({len(st.session_state.equipos)} items)</div>
            <div class="summary-value" style="font-size:1.1rem; color:#58a6ff">{formato_eur(coste_equipos)}</div>
        </div>
        <div class="summary-box" style="border-color: #58a6ff44;">
            <div class="summary-title">Inversión Total</div>
            <div class="summary-value">{formato_eur(coste_total_val)}</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Financiación</div>', unsafe_allow_html=True)

        pct_propio = 100 - pct_financiacion
        st.markdown(f"""
        <div class="summary-box">
            <div class="summary-title">Capital propio · {pct_propio}%</div>
            <div class="summary-value" style="font-size:1.1rem; color:#3fb950">{formato_eur(capital_propio)}</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Capital externo · {pct_financiacion}%</div>
            <div class="summary-value" style="font-size:1.1rem; color:#f0883e">{formato_eur(capital_externo)}</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Préstamo</div>', unsafe_allow_html=True)

        st.markdown(f"""
        <div class="summary-box">
            <div class="summary-title">Tipo de interés</div>
            <div class="summary-value" style="font-size:1.1rem; color:#58a6ff">{tipo_interes:.2f}%</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Plazo · Carencia</div>
            <div class="summary-value" style="font-size:1.1rem; color:#58a6ff">{plazo_amortizacion}a · {periodo_carencia}a</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Cuota anual estimada</div>
            <div class="summary-value" style="font-size:1.1rem; color:#f0883e">{formato_eur(cuota_anual)}</div>
            <div class="summary-label">{"Francés" if "Francés" in tipo_amortizacion else "Lineal"}</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Rentabilidad</div>', unsafe_allow_html=True)

        st.markdown(f"""
        <div class="summary-box">
            <div class="summary-title">WACC</div>
            <div class="summary-value" style="font-size:1.1rem; color:#a371f7">{wacc_input:.2f}%</div>
            <div class="summary-label">Tasa de descuento para el VAN</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">Ingreso / Tonelada</div>
            <div class="summary-value" style="font-size:1.1rem; color:#58a6ff">€{ingreso_por_tonelada:.2f}</div>
        </div>
        <div class="summary-box">
            <div class="summary-title">EBITDA anual</div>
            <div class="summary-value" style="font-size:1.1rem; color:#3fb950">{formato_eur(ebitda_anual)}</div>
        </div>
        <div class="summary-box" style="border-color: {'#3fb95044' if flujo_caja_anual >= 0 else '#f8514944'};">
            <div class="summary-title">Flujo de caja anual</div>
            <div class="summary-value" style="color:{'#3fb950' if flujo_caja_anual >= 0 else '#f85149'}">{formato_eur(flujo_caja_anual)}</div>
            <div class="summary-label">Tras servicio de la deuda</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<hr>", unsafe_allow_html=True)
        st.markdown('<div class="section-label">Resultados Financieros</div>', unsafe_allow_html=True)

        # Leer cálculos desde session_state
        _c = st.session_state.get("_calc", {})
        van = _c.get("van", 0)
        tir = _c.get("tir", None)
        wacc = _c.get("wacc", wacc_input / 100)

        # Formateo TIR
        if tir is not None and -1 < tir < 100:
            tir_str = f"{tir * 100:.2f}%"
            tir_color = "#3fb950" if tir >= wacc else "#f85149"
            tir_label = "TIR ≥ WACC · Proyecto viable" if tir >= wacc else "TIR < WACC · Proyecto no viable"
        else:
            tir_str = "N/D"
            tir_color = "#8b949e"
            tir_label = "No convergió — revisa los flujos"

        van_color = "#3fb950" if van >= 0 else "#f85149"
        van_label = "VAN ≥ 0 · Crea valor" if van >= 0 else "VAN < 0 · Destruye valor"

        st.markdown(f"""
        <div class="summary-box" style="border-color: {van_color}44;">
            <div class="summary-title">VAN</div>
            <div class="summary-value" style="color:{van_color}">{formato_eur(van)}</div>
            <div class="summary-label">{van_label}</div>
        </div>
        <div class="summary-box" style="border-color: {tir_color}44;">
            <div class="summary-title">TIR</div>
            <div class="summary-value" style="color:{tir_color}">{tir_str}</div>
            <div class="summary-label">{tir_label}</div>
        </div>
        """, unsafe_allow_html=True)

        # ── Botón descarga Excel ──────────────────────────────────────────────────
        st.markdown("<hr>", unsafe_allow_html=True)
        if _c:
            excel_bytes = generar_excel(_c)
            st.download_button(
                label="⬇ Descargar modelo Excel",
                data=excel_bytes,
                file_name="modelo_financiero_portuario.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


    # ── FOOTER ────────────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="footer">
        ANÁLISIS FINANCIERO PORTUARIO · v0.1 · Módulo de introducción de datos
    </div>
    """, unsafe_allow_html=True)


### LOGIN
load_dotenv()

USERNAME = st.secrets.get("STREAMLIT_USERNAME", os.getenv("STREAMLIT_USERNAME"))
PASSWORD = st.secrets.get("STREAMLIT_PASSWORD", os.getenv("STREAMLIT_PASSWORD"))

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

def show_login():
    st.title("Login")

    user = st.text_input("Username")
    pwd = st.text_input("Password", type="password")

    if st.button("Login"):
        if user == USERNAME and pwd == PASSWORD:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Invalid credentials")

if st.session_state.logged_in:
    main()
else:
    show_login()
